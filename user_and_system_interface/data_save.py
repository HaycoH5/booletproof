import os
from datetime import datetime
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


class DataSave:
    """Сохраняет данные формата JSON в текстовые и табличные форматы."""

    def __init__(self, base_dir: str, base_table: str):
        """Инициализация объекта сохранения данных."""
        self.base_dir = base_dir
        self.current_table_name = base_table

        self.headers = [
            "Дата", "Подразделение", "Операция", "Культура",
            "За день, га", "Начала операции", "Вал за день, ц", "Вал с начала, ц"
        ]

        os.makedirs(base_dir, exist_ok=True)

    def current_data(self) -> str:
        """Возвращает текущую дату и время в формате: ЧасыДеньМесяцГод."""
        return datetime.now().strftime("%H%d%m%Y")

    def convert_iso_to_custom_format(self, iso_string: str) -> str:
        """Преобразует ISO-дату в формат: Минута_Час_День_Месяц_Год."""
        dt = datetime.strptime(iso_string, "%Y-%m-%dT%H:%M:%S.%fZ")
        return dt.strftime("%M_%H_%d_%m_%Y")

    def _get_next_message_number(self, sender: str) -> int:
        """Возвращает следующий порядковый номер сообщения от отправителя."""
        return sum(sender in filename for filename in os.listdir(self.base_dir)) + 1

    def save_to_txt(self, data: Dict[str, Any]) -> None:
        """Сохраняет сообщение в текстовом формате."""
        sender = data["from"]
        timestamp = data["timestamp"]
        content = data["content"]

        message_num = self._get_next_message_number(sender)
        timestamp_str = self.convert_iso_to_custom_format(timestamp)
        file_name = f"{sender}_{message_num}_{timestamp_str}.txt"
        file_path = os.path.join(self.base_dir, file_name)

        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content + '\n')

    def append_message_to_table(
        self, filepath: str, message_dict: Dict[str, Any], date_value: str
    ) -> None:
        """
        Добавляет строку в Excel-таблицу. Пустые значения, включая автозаполненную дату, выделяются жёлтым.
        """
        full_path = os.path.join(filepath, self.current_table_name)
        wb = load_workbook(full_path)
        ws = wb.active

        headers = [
            "Дата", "Подразделение", "Операция", "Культура",
            "За день, га", "С начала операции, га",
            "Вал за день, ц", "Вал с начала, ц", "Исходное сообщение"
        ]

        normalized_headers = [h.strip().lower() for h in headers]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        header_row = None
        header_map = {}

        # Поиск строки с заголовками
        for row in ws.iter_rows(min_row=1, max_row=50):
            values = [str(cell.value).strip().lower() if cell.value else '' for cell in row]
            match_count = sum(1 for val in values if val in normalized_headers)

            if match_count >= len(headers) - 2:
                header_row = row[0].row
                for i, val in enumerate(values):
                    if val in normalized_headers:
                        header_map[val] = i + 1
                break

        if not header_row:
            raise ValueError("Не найдена строка с заголовками.")

        # Поиск первой пустой строки
        next_row = header_row + 1
        while ws.cell(row=next_row, column=header_map.get("подразделение", 2)).value:
            next_row += 1
        
        # Заполнение новой строки
        for header in headers:
            key = header.strip().lower()
            col = header_map.get(key)

            if not col:
                continue

            value = message_dict.get(header, "")
            fill = None

            if header == "Дата" and not value:
                value = date_value
                fill = yellow_fill
            elif value in [None, ""]:
                fill = yellow_fill

            cell = ws.cell(row=next_row, column=col, value=value)
            if fill:
                cell.fill = fill

        new_table_name = full_path.replace("template.xlsx", f"{self.current_data()}_BulletProof.xlsx")
        wb.save(new_table_name)

        self.current_table_name = new_table_name
