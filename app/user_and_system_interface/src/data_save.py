import os
from datetime import datetime
from typing import Dict, Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, PatternFill, Alignment

class DataSave:
    """Сохраняет данные формата JSON в текстовые и табличные форматы."""

    def __init__(self, base_dir: str, base_table: str):
        """Инициализация объекта сохранения данных."""
        self.base_dir = base_dir
        self.current_table_name = base_table

        self.headers = [
            "Дата", "Подразделение", "Операция", "Культура",
            "За день, га", "Начала операции", "Вал за день, ц", "Вал с начала, ц"
        ]

        os.makedirs(base_dir, exist_ok=True)

    def current_data(self) -> str:
        """Возвращает текущую дату и время в формате: ЧасыДеньМесяцГод."""
        return datetime.now().strftime("%H%d%m%Y")

    def convert_iso_to_custom_format(self, iso_string: str) -> str:
        """Преобразует ISO-дату в формат: Минута_Час_День_Месяц_Год."""
        dt = datetime.strptime(iso_string, "%Y-%m-%dT%H:%M:%S.%fZ")
        return dt.strftime("%M_%H_%d_%m_%Y")

    def _get_next_message_number(self, sender: str) -> int:
        """Возвращает следующий порядковый номер сообщения от отправителя."""
        return sum(sender in filename for filename in os.listdir(self.base_dir)) + 1

    def save_to_txt(self, data: Dict[str, Any], path: str) -> None:
        """Сохраняет сообщение в текстовом формате."""
        sender = data["from"]
        timestamp = data["timestamp"]
        content = data["content"]

        message_num = self._get_next_message_number(sender)
        timestamp_str = self.convert_iso_to_custom_format(timestamp)
        file_name = f"{sender}_{message_num}_{timestamp_str}.txt"
        file_path = path + "/" + file_name

        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content + '\n')


    def append_message_to_table(self, filepath: str, message_dict: Dict[str, Any], date_value: str) -> None:
        """
        Добавляет строку в Excel-таблицу с проверкой типов данных.
        Если значение не прошло проверку, оно сохраняется как есть и выделяется жёлтым.
        """
        for message in message_dict:
            if self.current_table_name not in os.listdir(filepath):
                self.create_agro_report(self.current_table_name, filepath)

            def cast_value(header: str, value: Any):
                try:
                    if header == "Дата":
                        return datetime.strptime(str(value), "%Y-%m-%d").date(), False
                    elif header in ["Подразделение", "Операция", "Культура"]:
                        return str(value), False
                    elif header in ["За день, га", "С начала операции, га", "Вал за день, ц", "Вал с начала, ц"]:
                        return int(float(value)), False
                except (ValueError, TypeError):
                    pass
                return value, True  # Вернуть исходное значение и флаг ошибки


            filepath_name = filepath + "/" + self.current_table_name
            wb = load_workbook(filepath_name)
            ws = wb.active

            headers = [
                "Дата", "Подразделение", "Операция", "Культура",
                "За день, га", "С начала операции, га",
                "Вал за день, ц", "Вал с начала, ц", "Исходное сообщение"
            ]

            normalized_headers = [h.strip().lower() for h in headers]
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            header_row = None
            header_map = {}

            # Поиск строки с заголовками
            for row in ws.iter_rows(min_row=1, max_row=50):
                values = [str(cell.value).strip().lower() if cell.value else '' for cell in row]
                match_count = sum(1 for val in values if val in normalized_headers)
                if match_count >= len(headers) - 2:
                    header_row = row[0].row
                    for i, val in enumerate(values):
                        if val in normalized_headers:
                            header_map[val] = i + 1
                    break

            if not header_row:
                raise ValueError("Не найдена строка с заголовками.")

            # Поиск первой пустой строки
            # Поиск первой пустой строки по первому столбцу (или более надёжно по "Дата")
            date_col = header_map.get("дата", 1)
            next_row = header_row + 1

            while ws.cell(row=next_row, column=date_col).value:
                next_row += 1

            # Заполнение новой строки
            for header in headers:
                key = header.strip().lower()
                col = header_map.get(key)
                if not col:
                    continue

                value = message.get(header, "")
                if header == "Дата" and not value:
                    value = date_value

                casted_value, error = cast_value(header, value)
                cell = ws.cell(row=next_row, column=col, value=casted_value)

                if casted_value in [None, ""] or error:
                    cell.fill = yellow_fill

            wb.save(filepath_name)

            # Переименование файла
            new_table_name = f"{self.current_data()}_BulletProof.xlsx"
            new_full_path = os.path.join(filepath, new_table_name)
            os.rename(filepath_name, new_full_path)

            self.current_table_name = new_table_name

    def create_agro_report(self, filename: str, path: str):
        # Создание книги и листа
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Заголовки таблицы
        headers = [
            "Дата", "Подразделение", "Операция", "Культура",
            "За день, га", "С начала операции, га", "Вал за день, ц", "Вал с начала, ц"
        ]

        # Стили
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        # Заполнение заголовков
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = 18

        # Добавляем пустые строки под таблицу
        for row in ws.iter_rows(min_row=3, max_row=22, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = center_align

        # Сохранение
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"

        path = path + "/" +  filename
        wb.save(path)
        self.current_table_name = filename

    def create_structure(self, paths_list):


        for path in paths_list:
            if not os.path.exists(path):
                os.mkdir(path)